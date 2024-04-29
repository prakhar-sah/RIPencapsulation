import openpyxl

# Define variable to load the dataframe
wb = openpyxl.load_workbook("../dumps/register_state_dump.xlsx")
 
# Define variable to read sheet
sheet = wb.worksheets[1]

ws3 = wb.create_sheet("possible_indirect_loads")
ws3_row = 0
ind_loads = [[''] * 2 for _ in range(200)]
 
# Iterate the loop to read the cell values
for row in range(2, sheet.max_row+1):
    if(16384 <= sheet.cell(row,1).value and sheet.cell(row,1).value < 24576):
        pc_diff = sheet.cell(row,1).value - sheet.cell(row-1,1).value
        if(0 < pc_diff <= 6):
            for i in range(4,16):
                reg_diff = sheet.cell(row,i).value - sheet.cell(row-1,i).value
                if(reg_diff != 0):
                    direct_load = 0
                    for j in range(1,16):
                        if(sheet.cell(row,i).value == sheet.cell(row-1,j).value):
                            direct_load = direct_load + 1
                    if(direct_load == 0):
                        ind_loads[ws3_row][0] = str(format(sheet.cell(row-1,1).value, '02X'))
                        ind_loads[ws3_row][1] = "R"+str(i-1)
                        ws3_row = ws3_row + 1
                    break;

for row in ind_loads:
    ws3.append(row)
    
unique_rows = set()
for row in ws3.iter_rows(1, values_only = True):
    if(row[0] != ""):
        unique_rows.add(tuple(row))

ws3.delete_rows(1, ws3.max_row)

for unique_row in unique_rows:
    ws3.append(unique_row)

# Save the Excel file
wb.save("../dumps/register_state_dump.xlsx")

